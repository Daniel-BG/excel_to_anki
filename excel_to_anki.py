import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
from pathlib import Path
import os

# Function to parse command-line arguments
def parse_args():
    parser = argparse.ArgumentParser(description="Generate Anki cards from an Excel file.")
    parser.add_argument('-i', '--input', required=True, help="Input Excel file (mandatory)")
    parser.add_argument('-o', '--output', default='anki_cards.tsv', help="Output TSV file (default: anki_cards.tsv)")
    parser.add_argument('-d', '--directory', default='anki_images', help="Output image directory (default: anki_images)")
    parser.add_argument('-c', '--column', type=int, required=True, help="Column number to take text from (mandatory)")
    parser.add_argument('-s', '--skip', type=int, default=1, help="How many rows to skip at the beginning (default: 1)")
    return parser.parse_args()

def main():
    args = parse_args()

    # Input and output paths
    excel_file = args.input
    output_tsv = args.output
    image_output_dir = args.directory
    text_col = args.column
    row_skip = args.skip

    # Create output directory for images
    if not os.path.exists(image_output_dir):
        os.makedirs(image_output_dir)

    # Load the workbook
    wb = load_workbook(excel_file)
    sheet = wb.active

    # Prepare data for the TSV file
    anki_data = []

    # Put your sheet in the loader
    image_loader = SheetImageLoader(sheet)

    # Iterate over rows
    for row in sheet.iter_rows(min_row=1+row_skip, values_only=False):  # Assuming the first row is a header
        text = row[text_col].value  # Adjust index if the text is in another column
        
        # Find and save first image in the row
        image_saved = False
        for cell in row:
            # Ask if there's an image in a cell
            if image_loader.image_in(cell.coordinate):
                image = image_loader.get(cell.coordinate)
                image_name = f'image_{text}.jpg'
                image_path = os.path.join(image_output_dir, image_name)
                image.save(image_path)
                anki_data.append([text, f'<img src="{image_name}">'])
                image_saved = True
                break
            
        if not image_saved:
            anki_data.append([text, ''])

    # Write the TSV file
    with open(output_tsv, 'w', encoding='utf-8') as tsv_file:
        for text, image_tag in anki_data:
            tsv_file.write(f'{text}\t{image_tag}\n')

    print(f"TSV file created at {output_tsv}. Images extracted to {image_output_dir}.")


# Entry point
if __name__ == "__main__":
    main()
